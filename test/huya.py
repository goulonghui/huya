# -*- coding:utf-8 -*-

# Created on 2023/5/10.
# -*- coding:utf-8 -*-

# Created on 2023/5/9.

import requests
import json
import time
import openpyxl
from io import BytesIO
from xlwt import Workbook
from typing import Iterable, Dict, Union, Tuple, Set
from urllib import parse

"""
https://udbsec.huya.com/web/appeal/launch
入参： {"uri":0,"data":{"user":"35184444002898"}}
返回值： {
    "uri": 0,
    "version": null,
    "context": null,
    "requestId": 0,
    "returnCode": 0,
    "message": null,
    "description": null,
    "data": {
        "result": 0,
        "user": "",
        "uid": 1199530126189,
        "sessionData": "AhUoY3YYjHgRK89AxASsMRWtf_lkHGFo0KNEntVfI1YG6ALF0-hbRJ7HSYqJNQBA_zmjh2fNPWjUUk4uLtXvX-59anfz2-qCJu_ESP-rNPnCgGoRe7ob6QPJHvkfmRrFEa9mTHha8F2vFTEVZ9Sqd20ZNGJFyMAI_Jesw4nPHmNdZjlL0PUXMO3BOqRszeq_RvIQzbA12Dz1La4V1XIiynTFhN6eLADFWkYsfp0Oy6xLJtKWUOrZiGKfC-i6qzV2gzqPL4UqLuhb8L2cTME5JhyeJaAA4WojfVKJ7Us4gKE7qxsa3rNiT5MOhxdf4zEkKg",
        "description": ""
    }
}

{
    "uri": 0,
    "version": null,
    "context": null,
    "requestId": 0,
    "returnCode": 90013,
    "message": "APPEAL_ORDER_IS_EXIST",
    "description": "此账号的申诉已存在",
    "data": null
}

{'uri': 1, 'version': None, 'context': None, 'requestId': 0, 'returnCode': 210003, 'message': '操作过于频繁，请稍后再试', 'description': '操作过于频繁，请稍后再试', 'data': None} <class 'dict'>
body: {'uri': 0, 'data': {'user': '35184537195586'}}
"""



headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.67 Safari/537.36"
}

data = {
    "uri": 0,
    "data": {
        "user": "35184444002898"
    }
}

def read_xls():
    # from openpyxl import load_workbook
    # # 1.打开 Excel 表格并获取表格名称
    # workbook = load_workbook(filename="huya.xlsx")
    # sheet = workbook["Sheet1"]
    #
    # for i in sheet.iter_rows(min_row=2):
    #     for j in i:
    #         print(j.value)
    return ["113550119"]

#
proxies = {
    'http': '183.162.226.250:40345',
    'https': '183.162.226.250:40345',
}


def request_post(_url, body):
    print("body:", body)
    # time.sleep(1)
    # proxies = {
    #     'http': 'http://egon:123@localhost:9743',
    #     'https': 'https://localhost:9743',
    # }
    response = requests.post(url=_url, json=body, proxies=proxies)
    print(response.status_code)
    print(response.text)
    return response


def request_get(_url):
    response = requests.get(url=_url)
    print(response.status_code)
    print(response.text)
    return response


def main():
    # response = request_get(proxy_url)
    # ip_list = []
    # resp = json.loads(response.text)
    # for _ in resp.get("data"):
    #     ip_list.append("%s:%s" % (_["ip"], _["port"]))

    for use_id in read_xls():
        body = {
            "uri": 0,
            "data": {
                "user": use_id
            }
        }
        response = request_post(url, body)
        if response.status_code != 200:
            pass
        resp = json.loads(response.text)
        print(resp, type(resp))
        if resp.get("returnCode") == 210003:
            pass
        if resp.get("message") == "APPEAL_ORDER_IS_EXIST":
            print("use_id: %s !!!!!!" % use_id)


if __name__ == '__main__':
    url = "https://udbsec.huya.com/web/appeal/launch"
    # proxy_url = "https://api.hailiangip.com:8522/api/getIpEncrypt?dataType=0&encryptParam=k60Va%2B4TTIoiEPFfhUCySBrmv%2F4utHOmExC44owkZhrdOA8y1i76cSUc0HJ2P6o5jWqgS6zERlFtF7BqxBumooSjSMhUAc59q1rbh7hrCLCVHIT%2By2AosUIaotYH3pnUZkHfjJt3bF%2BWYDKGVCI6sP9GciICyKc3M%2BgeMhyOoufYFsJvJYxeLU9YDql4IJq6vs3ERyFHZ9FAgNS8WBDIMt0Jv%2FQlqwlcd4gkrYI6AFg%3D"
    proxy_url = "http://ecs.hailiangip.com:8422/api/getIpEncrypt?dataType=0&encryptParam=MBftJikKr3D5USLE1py4JK%2FfT7GfQ3maxDbxprdoOkx%2FGQEB0bczI8I1Y8m2yCtLjWqgS6zERlFtF7BqxBumooSjSMhUAc59q1rbh7hrCLDTt453xipeXbywCAmTd8xv8zOQfaN025l1qdjvGzvJHUUmjnRyEndKEnFQFqOzQAhXwMMY2Pp7wRNtgRIJmPbHvs3ERyFHZ9FAgNS8WBDIMl7%2FeDXlL0x6IKTgy4kKtwD10%2FrggxuKwg%2Fa3uSVATqr"
    main()
    # read_xls()



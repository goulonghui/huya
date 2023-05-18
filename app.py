# -*- coding:utf-8 -*-

# Created on 2023/5/13.

import multiprocessing
import os
import time

from flask import Flask, request, render_template, redirect, url_for
from openpyxl import load_workbook

import crontab_task
from common import config
from common import db
from common import logger

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.mkdir(UPLOAD_FOLDER)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
crontab_task_process = None


@app.route('/')
def index():
    configs = db.get_all_config()
    return render_template('index.html', configs=configs)


@app.route('/edit/<int:config_id>', methods=['GET', 'POST'])
def edit_config(config_id):
    _config = db.get_config_by_id(config_id)
    if request.method == 'POST':
        db.update_config_by_id(config_id, request.form['key'], request.form['value'])
        if request.form['key'] == "task_interval_minutes":
            logger.log_info("config update. will restart crontab_task")
            logger.log_info(f"crontab_task_process: {crontab_task_process}")
            crontab_task_process.terminate()
            time.sleep(5)
            start_crontab()
        return redirect(url_for('index'))
    return render_template('edit.html', config=_config)


@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], file.filename))

            workbook = load_workbook(filename=os.path.join(app.config['UPLOAD_FOLDER'], file.filename))
            sheet = workbook["Sheet1"]
            users = []
            for row in sheet.iter_rows(min_row=2):
                for col in row:
                    users.append({"user_id": str(col.value).strip()})
            db.clear_user()
            db.bulk_create_user(users)
            return "批量替换账号id成功"
    return '批量替换账号id失败'


def run():
    logger.log_init("flask", config.log_config['log_dir'], "flask", config.log_config['log_level'])
    logger.log_info("server run on 0.0.0.0:5000")
    start_crontab()
    app.run(host='0.0.0.0', port=5000, debug=False)


def start_crontab():
    global crontab_task_process
    crontab_task_process = multiprocessing.Process(target=crontab_task.run)
    crontab_task_process.start()


if __name__ == '__main__':
    db.init_db()
    run()
